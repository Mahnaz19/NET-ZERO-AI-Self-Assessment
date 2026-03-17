import json
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - helper script
    raise SystemExit(
        "openpyxl is required to inspect the questionnaire. "
        "Install it with `pip install openpyxl` and re-run."
    ) from exc


def main() -> None:
    base = Path(__file__).resolve().parents[1]
    path = base / "data" / "raw" / "Minimum_Input_Questionnaire.xlsx"
    if not path.exists():
        raise SystemExit(
            "Missing questionnaire file at data/raw/Minimum_Input_Questionnaire.xlsx — "
            "please add it before running this prompt."
        )

    wb = load_workbook(path, data_only=True)
    if "FormSpec" not in wb.sheetnames:
        raise SystemExit(f'Expected sheet \"FormSpec\"; found {wb.sheetnames}')

    ws = wb["FormSpec"]

    # Header row
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    rows = []
    for row in ws.iter_rows(min_row=2):
        values = [cell.value for cell in row]
        if all(v is None for v in values):
            continue
        rows.append(values)

    print(
        json.dumps(
            {
                "headers": headers,
                "rows": rows,
            },
            indent=2,
            default=str,
        )
    )


if __name__ == "__main__":
    main()

